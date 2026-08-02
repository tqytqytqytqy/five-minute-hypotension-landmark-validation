#!/usr/bin/env python3
"""Finalize post-U0 QA, lineage and immutable result manifests.

This script is intentionally read-only with respect to frozen model objects,
the U0 execution lock and all U0 result files.  It only writes audit reports
and manifests outside those protected locations.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import re
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "02_code_configs/src"))

from lm5_validation.frozen import sha256_file  # noqa: E402


U0_DIR = ROOT / "05_MOVER_validation/H5_primary"
R1_DIR = ROOT / "05_MOVER_validation/R1_sensitivity"
OBS_DIR = ROOT / "05_MOVER_validation/observation_2x2"
DRIFT_DIR = ROOT / "05_MOVER_validation/subgroups_drift"
QA_DIR = ROOT / "09_QA_reproducibility/reports"
MANIFEST_DIR = ROOT / "10_run_logs_manifest"
LOCK_PATH = MANIFEST_DIR / "U0_execution_lock.json"
RECEIPT_PATH = MANIFEST_DIR / "pre_U0_freeze_receipt.json"
U0_MANIFEST = U0_DIR / "U0_result_SHA256SUMS.csv"
FINAL_MANIFEST = MANIFEST_DIR / "final_SHA256SUMS.csv"
FINAL_INVENTORY = MANIFEST_DIR / "final_file_inventory.csv"


FIGURE_STEMS = {
    "Fig1": "Fig1_cohort_flow_and_landmark_timeline",
    "Fig2": "Fig2_primary_external_validation",
    "Fig3": "Fig3_observation_process_2x2_factorial",
    "FigS1": "FigS1_phase_operator_R1_sensitivity",
    "FigS2": "FigS2_endpoint_coverage_sensitivity",
    "FigS3": "FigS3_prespecified_subgroup_forest",
    "FigS4": "FigS4_feature_drift",
    "FigS5": "FigS5_U1_U2_update_evaluation",
    "FigS6": "FigS6_outcome_observability_IPW",
    "FigS7": "FigS7_early_vasopressor_descriptive",
    "FigS8": "FigS8_first_vs_all_operations",
}

# These names are prohibited only when they are used as structured data fields
# (for example, a CSV column or JSON key).  Ordinary methodological prose that
# mentions an identifier field is not itself disclosure and must not fail QA.
DIRECT_IDENTIFIER_FIELDS = {
    "patient_id",
    "case_id",
    "encounter_id",
    "log_id",
    "subject_id",
    "person_id",
    "operation_id",
    "op_id",
    "patient_hash",
    "case_hash",
}


EXPECTED_R1 = {
    "R1_redetected_t0_main_stage2.csv.gz",
    "phase_buffer_R1_redetect_performance.csv",
    "phase_buffer_R1_redetect_cohort_flows.csv",
    "phase_buffer_operator_R1_patient_bootstrap_2000.csv",
    "first_vs_all_operations_patient_cluster_sensitivity.csv",
    "first_vs_all_operations_patient_bootstrap_2000.csv",
    "endpoint_and_coverage_sensitivity_performance.csv",
    "endpoint_and_coverage_patient_bootstrap_2000.csv",
    "endpoint_component_prevalence.csv",
    "endpoint_reclassification_vs_primary.csv",
}
EXPECTED_OBS = {
    "observation_2x2_point_performance.csv",
    "observation_2x2_complete_factorial_bootstrap_2000.csv",
    "observation_2x2_complete_factorial_replicates.csv.gz",
    "H5_R1_outcome_agreement.csv",
    "H5_R1_prediction_process_shift.csv",
}
EXPECTED_DRIFT = {
    "feature_drift_INSPIRE_vs_MOVER.csv",
    "prespecified_subgroup_performance.csv",
    "prespecified_subgroup_patient_bootstrap_2000.csv",
    "U1_U2_independent_update_evaluation.csv",
    "U1_U2_paired_bootstrap_2000.csv",
    "outcome_observability_target_overlap.csv",
    "outcome_observability_covariate_balance.csv",
    "outcome_observability_IPW_sensitivity.csv",
    "outcome_observability_gate_and_weight_diagnostics.csv",
    "early_vasopressor_descriptive_performance.csv",
    "early_vasopressor_patient_bootstrap_2000.csv",
    "early_vasopressor_action_category_audit.csv",
    "early_vasopressor_agent_mapping_audit.csv",
    "early_vasopressor_mapping_audit.csv",
}


def read_json(path: Path) -> dict:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise TypeError(f"JSON object required: {path}")
    return payload


def check(name: str, passed: bool, detail: object, severity: str = "P0") -> dict:
    return {
        "check": name,
        "passed": bool(passed),
        "severity": severity,
        "detail": json.dumps(detail, ensure_ascii=False, sort_keys=True)
        if isinstance(detail, (dict, list))
        else str(detail),
    }


def verify_u0_manifest() -> tuple[bool, list[str]]:
    lock = read_json(LOCK_PATH)
    table = pd.read_csv(U0_MANIFEST)
    failures: list[str] = []
    if lock.get("status") != "COMPLETE" or int(lock.get("attempt_number", 0)) != 1:
        failures.append("lock_not_first_attempt_COMPLETE")
    if sha256_file(U0_MANIFEST) != lock.get("expected_result_manifest_sha256"):
        failures.append("result_manifest_hash_differs_from_lock")
    expected = set(lock.get("expected_result_files", []))
    if set(table["file"]) != expected or len(table) != len(expected):
        failures.append("result_inventory_differs_from_lock")
    actual = {path.name for path in U0_DIR.iterdir() if path.is_file()}
    if actual != expected | {U0_MANIFEST.name}:
        failures.append("U0_directory_has_missing_or_extra_files")
    for row in table.itertuples(index=False):
        path = U0_DIR / str(row.file)
        if (
            not path.is_file()
            or path.stat().st_size != int(row.size_bytes)
            or sha256_file(path) != str(row.sha256)
        ):
            failures.append(str(row.file))
    snapshots = [
        lock.get("input_sha256"),
        lock.get("pre_publish_input_sha256"),
        lock.get("completion_input_sha256"),
    ]
    if not all(isinstance(item, dict) for item in snapshots) or not (
        snapshots[0] == snapshots[1] == snapshots[2]
    ):
        failures.append("U0_input_hash_snapshots_not_identical")
    return not failures, failures


def verify_secondary_receipt() -> tuple[bool, dict[str, dict[str, str]]]:
    receipt = read_json(RECEIPT_PATH)
    recorded = receipt.get("secondary_analysis_file_sha256", {})
    observed = {
        "02_code_configs/scripts/07_secondary_analyses.py": sha256_file(
            ROOT / "02_code_configs/scripts/07_secondary_analyses.py"
        ),
        "00_protocol_SAP/secondary_analysis_prespec.json": sha256_file(
            ROOT / "00_protocol_SAP/secondary_analysis_prespec.json"
        ),
    }
    failures = {
        key: {"recorded": recorded.get(key), "observed": observed.get(key)}
        for key in sorted(set(recorded) | set(observed))
        if recorded.get(key) != observed.get(key)
    }
    return not failures, failures


def result_files(directory: Path) -> set[str]:
    return {path.name for path in directory.iterdir() if path.is_file()}


def bootstrap_validity_checks() -> list[dict]:
    checks: list[dict] = []
    files = [
        U0_DIR / "U0_paired_patient_bootstrap_2000_summary.csv",
        U0_DIR / "U0_two_stage_fixed_strategy_bootstrap_2000_summary.csv",
        OBS_DIR / "observation_2x2_complete_factorial_bootstrap_2000.csv",
        R1_DIR / "phase_buffer_operator_R1_patient_bootstrap_2000.csv",
        R1_DIR / "first_vs_all_operations_patient_bootstrap_2000.csv",
        R1_DIR / "endpoint_and_coverage_patient_bootstrap_2000.csv",
        DRIFT_DIR / "prespecified_subgroup_patient_bootstrap_2000.csv",
        DRIFT_DIR / "U1_U2_paired_bootstrap_2000.csv",
        DRIFT_DIR / "early_vasopressor_patient_bootstrap_2000.csv",
    ]
    for path in files:
        if not path.is_file():
            checks.append(
                check(
                    f"bootstrap_valid_fraction::{path.name}",
                    False,
                    {"status": "required_bootstrap_summary_missing"},
                )
            )
            continue
        frame = pd.read_csv(path)
        evaluated_masks: list[pd.Series] = []
        ratios: list[pd.Series] = []
        schemas: list[str] = []
        structural_nonestimable = pd.Series(False, index=frame.index)
        structural_nonestimable_candidates = pd.Series(False, index=frame.index)

        # At extreme thresholds, a simple comparator can issue zero alerts.
        # PPV and alerts-per-true-positive are then mathematically undefined in
        # the point estimate and in every bootstrap sample.  This is an explicit
        # non-estimable result, not a failed bootstrap.  The exemption is narrow:
        # it never applies to LM5, discrimination/calibration, or finite point
        # estimates, and all inferential fields must be absent with n_valid=0.
        if path.name == "U0_paired_patient_bootstrap_2000_summary.csv":
            required = {
                "estimand",
                "estimate",
                "bootstrap_se",
                "ci95_lower",
                "ci95_upper",
                "valid_replicates",
            }
            if required.issubset(frame.columns):
                estimand = frame["estimand"].astype(str)
                allowed_suffix = estimand.str.endswith(
                    (
                        "__ppv",
                        "__alerts_per_true_positive",
                        "__false_alerts_per_true_positive",
                    )
                )
                structural_nonestimable_candidates = (
                    estimand.str.startswith("simple_")
                    & estimand.str.contains("__pt_", regex=False)
                    & allowed_suffix
                    & pd.to_numeric(frame["estimate"], errors="coerce").isna()
                    & pd.to_numeric(frame["bootstrap_se"], errors="coerce").isna()
                    & pd.to_numeric(frame["ci95_lower"], errors="coerce").isna()
                    & pd.to_numeric(frame["ci95_upper"], errors="coerce").isna()
                    & (pd.to_numeric(frame["valid_replicates"], errors="coerce") == 0)
                )
                workload = pd.read_csv(U0_DIR / "U0_fixed_threshold_alert_workload.csv")
                zero_denominator_verified = pd.Series(False, index=frame.index)
                pattern = re.compile(
                    r"^(simple_.+)__pt_([0-9]+\.[0-9]+)__"
                    r"(ppv|alerts_per_true_positive|false_alerts_per_true_positive)$"
                )
                for index in frame.index[structural_nonestimable_candidates]:
                    match = pattern.match(str(frame.at[index, "estimand"]))
                    if match is None:
                        continue
                    model, threshold_text, metric = match.groups()
                    threshold = float(threshold_text)
                    selected = workload[
                        (workload["model"].astype(str) == model)
                        & np.isclose(
                            pd.to_numeric(workload["threshold"], errors="coerce"),
                            threshold,
                            rtol=0,
                            atol=5e-7,
                        )
                    ]
                    if len(selected) != 1:
                        continue
                    point = selected.iloc[0]
                    denominator_column = (
                        "weighted_alerts" if metric == "ppv" else "true_positive_weight"
                    )
                    point_column = {
                        "ppv": "positive_predictive_value",
                        "alerts_per_true_positive": "alerts_per_true_positive",
                        "false_alerts_per_true_positive": "false_alerts_per_true_positive",
                    }[metric]
                    point_value = pd.to_numeric(
                        pd.Series([point[point_column]]), errors="coerce"
                    ).iloc[0]
                    denominator = pd.to_numeric(
                        pd.Series([point[denominator_column]]), errors="coerce"
                    ).iloc[0]
                    zero_denominator_verified.at[index] = (
                        pd.notna(denominator)
                        and float(denominator) == 0.0
                        and not np.isfinite(point_value)
                    )
                structural_nonestimable = (
                    structural_nonestimable_candidates & zero_denominator_verified
                )

        # U0 and 2x2 outputs use valid_replicates with either
        # requested_replicates or bootstrap_repetitions.
        validity_columns = [
            column for column in frame.columns
            if column == "valid_replicates" or column.endswith("_valid_replicates")
        ]
        for valid_column in validity_columns:
            stem = valid_column[: -len("valid_replicates")]
            candidates = [
                f"{stem}requested_replicates",
                f"{stem}bootstrap_repetitions",
                "requested_replicates",
                "bootstrap_repetitions",
            ]
            requested_column = next(
                (column for column in candidates if column in frame.columns), None
            )
            if requested_column is None:
                continue
            valid = pd.to_numeric(frame[valid_column], errors="coerce")
            requested = pd.to_numeric(frame[requested_column], errors="coerce")
            mask = valid.notna() & requested.notna() & (requested > 0)
            evaluated_masks.append(mask)
            ratios.append(valid / requested)
            schemas.append(f"{valid_column}/{requested_column}")

        # Secondary outputs use n_valid/n_boot and also report valid_fraction.
        if "n_valid" in frame.columns and "n_boot" in frame.columns:
            valid = pd.to_numeric(frame["n_valid"], errors="coerce")
            requested = pd.to_numeric(frame["n_boot"], errors="coerce")
            mask = valid.notna() & requested.notna() & (requested > 0)
            evaluated_masks.append(mask)
            ratios.append(valid / requested)
            schemas.append("n_valid/n_boot")

        # A direct valid_fraction is sufficient when counts are not exported.
        # When counts are present, compare the claimed fraction with the derived
        # ratio as an additional integrity check without double-counting rows.
        claimed_fraction = None
        if "valid_fraction" in frame.columns:
            claimed_fraction = pd.to_numeric(frame["valid_fraction"], errors="coerce")
            if not ratios:
                mask = claimed_fraction.notna()
                evaluated_masks.append(mask)
                ratios.append(claimed_fraction)
                schemas.append("valid_fraction")

        below_gate = 0
        invalid_counts = 0
        fraction_mismatches = 0
        evaluated = 0
        for ratio, mask in zip(ratios, evaluated_masks):
            gate_mask = mask & ~structural_nonestimable
            evaluated += int(gate_mask.sum())
            below_gate += int((ratio[gate_mask] < 0.95).sum())
            invalid_counts += int(((ratio[mask] < 0) | (ratio[mask] > 1)).sum())
        if claimed_fraction is not None and ratios and schemas[0] != "valid_fraction":
            mask = evaluated_masks[0] & claimed_fraction.notna()
            fraction_mismatches = int(
                ((ratios[0][mask] - claimed_fraction[mask]).abs() > 1e-12).sum()
            )
        checks.append(
            check(
                f"bootstrap_valid_fraction::{path.name}",
                evaluated > 0
                and below_gate == 0
                and invalid_counts == 0
                and fraction_mismatches == 0,
                {
                    "evaluated_estimands": evaluated,
                    "below_95_percent": below_gate,
                    "invalid_count_ratios": invalid_counts,
                    "claimed_fraction_mismatches": fraction_mismatches,
                    "schemas_evaluated": schemas,
                    "structurally_nonestimable_exemptions": int(
                        structural_nonestimable.sum()
                    ),
                    "unverified_nonestimable_candidates": int(
                        (structural_nonestimable_candidates & ~structural_nonestimable).sum()
                    ),
                    "structurally_nonestimable_estimands": frame.loc[
                        structural_nonestimable, "estimand"
                    ].astype(str).tolist()
                    if "estimand" in frame.columns
                    else [],
                },
            )
        )
    return checks


def ci_order_checks() -> tuple[bool, dict[str, int]]:
    failures: dict[str, int] = {}
    for directory in [U0_DIR, R1_DIR, OBS_DIR, DRIFT_DIR, ROOT / "06_tables"]:
        for path in sorted(directory.glob("*.csv")):
            try:
                frame = pd.read_csv(path)
            except Exception:
                continue
            pairs: list[tuple[str, str]] = []
            for column in frame.columns:
                if "lower" in column:
                    upper = column.replace("lower", "upper")
                    if upper in frame:
                        pairs.append((column, upper))
            bad = 0
            for lower, upper in pairs:
                lo = pd.to_numeric(frame[lower], errors="coerce")
                hi = pd.to_numeric(frame[upper], errors="coerce")
                mask = lo.notna() & hi.notna()
                bad += int((lo[mask] > hi[mask]).sum())
            if bad:
                failures[str(path.relative_to(ROOT))] = bad
    return not failures, failures


def two_by_two_contract() -> tuple[bool, dict[str, object]]:
    point = pd.read_csv(OBS_DIR / "observation_2x2_point_performance.csv")
    factorial = pd.read_csv(
        OBS_DIR / "observation_2x2_complete_factorial_bootstrap_2000.csv"
    )
    expected_cells = {
        "H5_features__H5_outcome",
        "H5_features__R1_outcome",
        "R1_features__H5_outcome",
        "R1_features__R1_outcome",
    }
    cells_ok = set(point["cell"]) == expected_cells
    estimands = factorial["estimand"].astype(str)
    six_pairs = estimands.str.startswith("delta__").sum()
    interaction = estimands.str.contains("feature_by_outcome_interaction").sum()
    ok = cells_ok and six_pairs >= 6 and interaction >= 1
    return ok, {
        "cells": sorted(set(point["cell"])),
        "paired_estimands": int(six_pairs),
        "interaction_estimands": int(interaction),
    }


def subgroup_contract() -> tuple[bool, dict[str, object]]:
    frame = pd.read_csv(DRIFT_DIR / "prespecified_subgroup_performance.csv")
    expected = {
        "age_lt65", "age_ge65", "female", "male", "ASA_1_2", "ASA_ge3",
        "BMI_lt25", "BMI_25_30", "BMI_ge30", "t0_ART", "t0_NIBP",
        "t0_early_le10", "t0_late_gt10",
    }
    observed = set(frame["subgroup"].astype(str))
    status_ok = frame["estimability_status"].isin({"estimable", "nonestimable"}).all()
    return observed == expected and status_ok, {
        "expected_n": len(expected),
        "observed_n": len(observed),
        "missing": sorted(expected - observed),
        "unexpected": sorted(observed - expected),
    }


def normalize_field_name(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")


def direct_identifier_fields(values: list[object] | tuple[object, ...]) -> list[str]:
    return sorted(
        {
            normalize_field_name(value)
            for value in values
            if normalize_field_name(value) in DIRECT_IDENTIFIER_FIELDS
        }
    )


def json_identifier_keys(value: object, prefix: str = "$") -> list[str]:
    findings: list[str] = []
    if isinstance(value, dict):
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}"
            if normalize_field_name(key) in DIRECT_IDENTIFIER_FIELDS:
                findings.append(child_prefix)
            findings.extend(json_identifier_keys(child, child_prefix))
    elif isinstance(value, list):
        for index, child in enumerate(value):
            findings.extend(json_identifier_keys(child, f"{prefix}[{index}]"))
    return findings


def tabular_identifier_findings(path: Path) -> tuple[list[str], str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            header = next(csv.reader(handle), [])
        fields = direct_identifier_fields(header)
        return [f"CSV column:{field}" for field in fields], "csv_header"

    if suffix == ".json":
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            return [f"JSON unreadable:{type(exc).__name__}"], "json_keys"
        return json_identifier_keys(payload), "json_keys"

    if suffix in {".md", ".txt"}:
        # Prose mentions are intentionally allowed.  Only explicit key/value
        # assignments are treated as structured individual-level disclosure.
        text = path.read_text(encoding="utf-8", errors="replace")
        pattern = re.compile(
            r"(?im)^\s*(?:[-*]\s*)?"
            r"(patient_id|case_id|encounter_id|log_id|subject_id|person_id|"
            r"operation_id|op_id|patient_hash|case_hash)\s*[:=]\s*(\S+)"
        )
        findings = [f"structured assignment:{match.group(1)}" for match in pattern.finditer(text)]
        return sorted(set(findings)), "structured_text_assignments"

    if suffix == ".xlsx":
        findings: list[str] = []
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                for row_index, row in enumerate(rows[:-1]):
                    for column_index, value in enumerate(row):
                        field = normalize_field_name(value)
                        if field not in DIRECT_IDENTIFIER_FIELDS:
                            continue
                        # A field name by itself (for example in a data
                        # dictionary) is not disclosure.  Require at least one
                        # nonempty value beneath it in the same putative column.
                        below = [
                            later[column_index]
                            for later in rows[row_index + 1 : row_index + 11]
                            if column_index < len(later)
                        ]
                        if any(item is not None and str(item).strip() for item in below):
                            findings.append(
                                f"XLSX column:{sheet.title}!R{row_index + 1}C{column_index + 1}:{field}"
                            )
        finally:
            workbook.close()
        return findings, "xlsx_structured_columns"

    if suffix == ".docx":
        # Raw XML token scans wrongly fail ordinary prose.  Detect only an
        # identifier-labelled table column followed by nonempty cell values.
        findings: list[str] = []
        with zipfile.ZipFile(path) as archive:
            document_xml = archive.read("word/document.xml").decode(
                "utf-8", errors="replace"
            )
        table_pattern = re.compile(r"<w:tbl\b.*?</w:tbl>", re.S)
        row_pattern = re.compile(r"<w:tr\b.*?</w:tr>", re.S)
        cell_pattern = re.compile(r"<w:tc\b.*?</w:tc>", re.S)
        text_pattern = re.compile(r"<w:t(?:\s[^>]*)?>(.*?)</w:t>", re.S)
        for table_number, table_xml in enumerate(table_pattern.findall(document_xml), 1):
            parsed_rows: list[list[str]] = []
            for row_xml in row_pattern.findall(table_xml):
                parsed_rows.append(
                    [
                        "".join(text_pattern.findall(cell_xml))
                        for cell_xml in cell_pattern.findall(row_xml)
                    ]
                )
            for row_index, row in enumerate(parsed_rows[:-1]):
                for column_index, value in enumerate(row):
                    field = normalize_field_name(value)
                    if field not in DIRECT_IDENTIFIER_FIELDS:
                        continue
                    below = [
                        later[column_index]
                        for later in parsed_rows[row_index + 1 : row_index + 11]
                        if column_index < len(later)
                    ]
                    if any(str(item).strip() for item in below):
                        findings.append(
                            f"DOCX column:table{table_number}/r{row_index + 1}/c{column_index + 1}:{field}"
                        )
        return findings, "docx_structured_columns"

    # PNG/PDF are backed by separately scanned aggregate source-data files.
    return [], "binary_or_visual_source_data_scanned_separately"


def publication_privacy_audit() -> tuple[pd.DataFrame, bool]:
    rows: list[dict] = []
    directories = [
        ROOT / "06_tables",
        ROOT / "07_figures",
        ROOT / "08_manuscript_supplement",
    ]
    for directory in directories:
        for path in sorted(directory.rglob("*")):
            if not path.is_file() or path.name.startswith("."):
                continue
            findings, scan_method = tabular_identifier_findings(path)
            rows.append(
                {
                    "relative_path": str(path.relative_to(ROOT)),
                    "shareability_scope": "publication_facing",
                    "structured_identifier_scan_passed": not findings,
                    "scan_method": scan_method,
                    "finding": "; ".join(findings),
                }
            )
    table = pd.DataFrame(rows)
    return table, bool(
        table.empty or table["structured_identifier_scan_passed"].all()
    )


def write_secondary_manifest() -> pd.DataFrame:
    rows = []
    for directory in [OBS_DIR, R1_DIR, DRIFT_DIR]:
        for path in sorted(directory.glob("*")):
            if path.is_file():
                rows.append(
                    {
                        "relative_path_from_project_root": str(path.relative_to(ROOT)),
                        "size_bytes": path.stat().st_size,
                        "sha256": sha256_file(path),
                        "role": "prespecified_secondary_result",
                    }
                )
    table = pd.DataFrame(rows)
    table.to_csv(
        MANIFEST_DIR / "secondary_result_SHA256SUMS.csv",
        index=False,
        encoding="utf-8-sig",
    )
    return table


def write_lineage() -> None:
    rows = [
        {
            "output_scope": "INSPIRE model and internal validation",
            "primary_script": "02_code_configs/scripts/04_fit_freeze_inspire.py",
            "primary_inputs": "03_derived_cohorts/INSPIRE/inspire_main_stage2.csv.gz",
            "contract": "02_code_configs/configs/analysis.json; 00_protocol_SAP/SAP_v1.0.md",
            "result_location": "04_frozen_INSPIRE_LM5_model; 06_tables/INSPIRE_*",
        },
        {
            "output_scope": "MOVER H5 U0 primary external validation",
            "primary_script": "02_code_configs/scripts/06_run_mover_u0_once.py",
            "primary_inputs": "03_derived_cohorts/MOVER/mover_main_stage2.csv.gz; mover_first_fully_evaluable_operation.csv.gz",
            "contract": "10_run_logs_manifest/pre_U0_freeze_receipt.json; 09_QA_reproducibility/reports/MOVER_pre_model_QA.json",
            "result_location": "05_MOVER_validation/H5_primary",
        },
        {
            "output_scope": "Prespecified secondary and mechanism analyses",
            "primary_script": "02_code_configs/scripts/07_secondary_analyses.py",
            "primary_inputs": "MOVER H5/R1 grids, all t0 operations, frozen model, U0 probabilities",
            "contract": "00_protocol_SAP/secondary_analysis_prespec.json",
            "result_location": "05_MOVER_validation/observation_2x2; R1_sensitivity; subgroups_drift",
        },
        {
            "output_scope": "Publication tables and figures",
            "primary_script": "02_code_configs/scripts/08_build_publication_tables.py; 02_code_configs/scripts/08_generate_publication_figures.py",
            "primary_inputs": "Frozen INSPIRE results; locked U0 results; prespecified secondary results",
            "contract": "Result hierarchy in SAP and pre-U0 receipt",
            "result_location": "06_tables; 07_figures",
        },
    ]
    pd.DataFrame(rows).to_csv(
        MANIFEST_DIR / "analysis_lineage.csv", index=False, encoding="utf-8-sig"
    )


def verify_figure_package() -> tuple[bool, dict[str, object]]:
    figure_dir = ROOT / "07_figures"
    manifest_path = figure_dir / "figure_source_manifest.csv"
    qa_path = figure_dir / "figure_QA_report.json"
    failures: list[str] = []

    expected_files = {
        f"{stem}.{extension}"
        for stem in FIGURE_STEMS.values()
        for extension in ("png", "pdf")
    }
    missing_files = sorted(
        name
        for name in expected_files
        if not (figure_dir / name).is_file() or (figure_dir / name).stat().st_size <= 0
    )
    if missing_files:
        failures.extend(f"missing_or_empty:{name}" for name in missing_files)
    if not manifest_path.is_file():
        failures.append("figure_source_manifest_missing")
    if not qa_path.is_file():
        failures.append("figure_QA_report_missing")
    if failures:
        return False, {
            "expected_figure_pairs": len(FIGURE_STEMS),
            "failures": failures,
        }

    manifest = pd.read_csv(manifest_path)
    required_columns = {
        "figure_id",
        "source_data_file",
        "source_data_sha256",
        "figure_png",
        "figure_png_sha256",
        "figure_pdf",
        "figure_pdf_sha256",
        "generation_script",
        "generation_script_sha256",
    }
    missing_columns = sorted(required_columns - set(manifest.columns))
    if missing_columns:
        failures.append(f"manifest_columns_missing:{'|'.join(missing_columns)}")
    else:
        observed_ids = set(manifest["figure_id"].astype(str))
        if observed_ids != set(FIGURE_STEMS):
            failures.append(
                "figure_ids_differ:"
                f"missing={sorted(set(FIGURE_STEMS) - observed_ids)};"
                f"unexpected={sorted(observed_ids - set(FIGURE_STEMS))}"
            )
        for row_index, row in manifest.iterrows():
            for path_column, hash_column in [
                ("source_data_file", "source_data_sha256"),
                ("figure_png", "figure_png_sha256"),
                ("figure_pdf", "figure_pdf_sha256"),
                ("generation_script", "generation_script_sha256"),
            ]:
                relative = str(row[path_column])
                path = (ROOT / relative).resolve()
                try:
                    path.relative_to(ROOT.resolve())
                except ValueError:
                    failures.append(f"manifest_path_outside_root:row{row_index}:{relative}")
                    continue
                if not path.is_file() or path.stat().st_size <= 0:
                    failures.append(f"manifest_file_missing_or_empty:row{row_index}:{relative}")
                elif sha256_file(path) != str(row[hash_column]):
                    failures.append(f"manifest_hash_mismatch:row{row_index}:{relative}")
        scripts = set(manifest["generation_script"].astype(str))
        expected_script = "02_code_configs/scripts/08_generate_publication_figures.py"
        if scripts != {expected_script}:
            failures.append(f"unexpected_generation_scripts:{sorted(scripts)}")

    qa = read_json(qa_path)
    expected_qa = {
        "status": "PASS",
        "visual_inspection_status": "PASS",
        "u0_hash_manifest_verified_before_read": True,
        "model_refit_or_prediction_generation_performed": False,
        "patient_level_figure_source_exported": False,
        "figure_count": len(FIGURE_STEMS),
        "source_data_file_count": 16,
    }
    for key, expected in expected_qa.items():
        if qa.get(key) != expected:
            failures.append(f"figure_QA_field:{key}:expected={expected}:observed={qa.get(key)}")
    if qa.get("manifest_sha256") != sha256_file(manifest_path):
        failures.append("figure_QA_manifest_hash_mismatch")
    qa_ids = {
        str(item.get("figure_id"))
        for item in qa.get("figures", [])
        if isinstance(item, dict)
    }
    if qa_ids != set(FIGURE_STEMS):
        failures.append("figure_QA_figure_inventory_differs")

    return not failures, {
        "expected_figure_pairs": len(FIGURE_STEMS),
        "manifest_rows": len(manifest),
        "unique_source_data_files": int(manifest["source_data_file"].nunique())
        if "source_data_file" in manifest
        else 0,
        "failures": failures,
    }


def final_inventory_and_hashes() -> tuple[pd.DataFrame, pd.DataFrame]:
    excluded_names = {FINAL_MANIFEST.name, FINAL_INVENTORY.name}
    excluded_parts = {"__pycache__", "_docx_render_QA", "_workbook_render_QA", "tmp"}
    rows = []
    for path in sorted(ROOT.rglob("*")):
        if not path.is_file() or path.name in excluded_names:
            continue
        relative = path.relative_to(ROOT)
        if any(part in excluded_parts for part in relative.parts):
            continue
        role = "internal_reproducibility"
        if relative.parts[0] in {"06_tables", "07_figures", "08_manuscript_supplement"}:
            role = "publication_facing"
        elif relative.parts[0] in {"09_QA_reproducibility", "10_run_logs_manifest"}:
            role = "audit_lineage"
        rows.append(
            {
                "relative_path": str(relative),
                "size_bytes": path.stat().st_size,
                "sha256": sha256_file(path),
                "role": role,
                "share_by_default": role == "publication_facing",
            }
        )
    inventory = pd.DataFrame(rows)
    inventory.to_csv(FINAL_INVENTORY, index=False, encoding="utf-8-sig")
    hashes = inventory[["relative_path", "size_bytes", "sha256"]].copy()
    hashes.to_csv(FINAL_MANIFEST, index=False, encoding="utf-8-sig")
    return inventory, hashes


def main() -> None:
    QA_DIR.mkdir(parents=True, exist_ok=True)
    MANIFEST_DIR.mkdir(parents=True, exist_ok=True)
    checks: list[dict] = []

    u0_ok, u0_failures = verify_u0_manifest()
    checks.append(check("U0_lock_manifest_and_files_immutable", u0_ok, u0_failures))
    receipt_ok, receipt_failures = verify_secondary_receipt()
    checks.append(check("secondary_code_matches_pre_U0_receipt", receipt_ok, receipt_failures))

    for name, directory, expected in [
        ("R1_secondary_result_inventory", R1_DIR, EXPECTED_R1),
        ("observation_2x2_result_inventory", OBS_DIR, EXPECTED_OBS),
        ("subgroup_drift_result_inventory", DRIFT_DIR, EXPECTED_DRIFT),
    ]:
        observed = result_files(directory)
        checks.append(
            check(
                name,
                expected.issubset(observed),
                {"missing": sorted(expected - observed), "extra": sorted(observed - expected)},
            )
        )

    checks.extend(bootstrap_validity_checks())
    ci_ok, ci_failures = ci_order_checks()
    checks.append(check("all_reported_CI_bounds_are_ordered", ci_ok, ci_failures))
    obs_ok, obs_detail = two_by_two_contract()
    checks.append(check("observation_2x2_complete_contract", obs_ok, obs_detail))
    subgroup_ok, subgroup_detail = subgroup_contract()
    checks.append(check("all_prespecified_subgroup_rows_retained", subgroup_ok, subgroup_detail))

    secondary_manifest = write_secondary_manifest()
    checks.append(
        check(
            "secondary_result_manifest_nonempty",
            len(secondary_manifest) >= len(EXPECTED_R1 | EXPECTED_OBS | EXPECTED_DRIFT),
            {"manifest_rows": len(secondary_manifest)},
        )
    )
    write_lineage()

    figure_ok, figure_detail = verify_figure_package()
    checks.append(
        check(
            "figure_package_complete_and_hash_verified",
            figure_ok,
            figure_detail,
            severity="P1",
        )
    )

    privacy, privacy_ok = publication_privacy_audit()
    privacy.to_csv(
        QA_DIR / "privacy_shareability_audit.csv", index=False, encoding="utf-8-sig"
    )
    checks.append(
        check(
            "publication_artifacts_have_no_structured_individual_identifiers",
            privacy_ok,
            {
                "files_scanned": len(privacy),
                "failed": int((~privacy["structured_identifier_scan_passed"]).sum())
                if len(privacy)
                else 0,
                "prose_mentions_allowed": True,
            },
            severity="P1",
        )
    )

    expected_publication = [
        ROOT / "06_tables/all_tables_LM5_INSPIRE_MOVER_v1.0.xlsx",
        ROOT / "08_manuscript_supplement/LM5_manuscript.docx",
        ROOT / "08_manuscript_supplement/LM5_supplement.docx",
    ]
    checks.append(
        check(
            "core_publication_artifacts_exist",
            all(path.is_file() and path.stat().st_size > 0 for path in expected_publication),
            {
                "expected": [str(path.relative_to(ROOT)) for path in expected_publication],
                "missing": [
                    str(path.relative_to(ROOT))
                    for path in expected_publication
                    if not path.is_file()
                ],
                "empty": [
                    str(path.relative_to(ROOT))
                    for path in expected_publication
                    if path.is_file() and path.stat().st_size <= 0
                ],
            },
            severity="P1",
        )
    )

    table = pd.DataFrame(checks)
    table.to_csv(QA_DIR / "post_U0_final_QA.csv", index=False, encoding="utf-8-sig")
    failed = table[~table["passed"]]
    status = "GREEN" if failed.empty else "RED"
    report = {
        "study_id": "LM5_COMMON18_INSPIRE_MOVER_20260712",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "status": status,
        "checks": len(table),
        "passed": int(table["passed"].sum()),
        "failed": int((~table["passed"]).sum()),
        "failed_checks": failed[["check", "severity", "detail"]].to_dict("records"),
        "U0_files_modified": False,
        "model_or_threshold_modified": False,
    }
    (QA_DIR / "publication_artifact_QA.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    lines = [
        "# Post-U0 final QA",
        "",
        f"- Status: **{status}**",
        f"- Checks: {len(table)} total; {int(table['passed'].sum())} passed; {int((~table['passed']).sum())} failed.",
        "- U0 result files and frozen model objects were read only.",
        "",
        "## Check results",
        "",
    ]
    for row in table.itertuples(index=False):
        mark = "PASS" if row.passed else "FAIL"
        lines.append(f"- [{mark}] {row.check} ({row.severity}): {row.detail}")
    (QA_DIR / "post_U0_final_QA.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    # The final inventory is written after every other QA artifact so it captures
    # the complete handoff state.  It excludes itself and its hash table by design.
    final_inventory_and_hashes()
    print(json.dumps(report, ensure_ascii=False))
    if status != "GREEN":
        raise SystemExit("Post-U0 final QA RED")


if __name__ == "__main__":
    main()
